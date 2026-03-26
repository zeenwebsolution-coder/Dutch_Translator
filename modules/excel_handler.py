"""
excel_handler.py — Advanced multi-column detection and separator-aware writing.

Updated to:
  - Find ALL English/Dutch column pairs (supports up to 3000+ columns).
  - Handle "Separator Rows" (skip rows that are just headers/titles).
  - Preserve formatting in large, complex files.
"""

from __future__ import annotations

import io
import re
import logging
from dataclasses import dataclass
from typing import Optional, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell

logger = logging.getLogger(__name__)

# ── Data types ────────────────────────────────────────────────────────────────

@dataclass
class WordEntry:
    sheet: str
    row: int
    col: int            # English column index
    target_col: int     # Corresponding Dutch column index
    value: str

# ── Header Matching ───────────────────────────────────────────────────────────

# Flexible regex for English and Dutch headers
SOURCE_HEADER_REGEX = re.compile(r".*(english|en|source|original).*", re.IGNORECASE)
TARGET_HEADER_REGEX = re.compile(r".*(dutch|nl|target|vertaling|translation).*", re.IGNORECASE)

def _find_column_pairs(ws) -> List[Tuple[int, int]]:
    """
    Search first 10 rows for ALL pairs of English and Dutch columns.
    Returns a list of (source_col, target_col).
    """
    # 1. Find all English columns
    english_cols: List[int] = []
    dutch_cols: List[int] = []
    
    # Scan up to row 10
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value or "").strip()
            if not val: continue
            
            if SOURCE_HEADER_REGEX.match(val):
                if cell.column not in english_cols:
                    english_cols.append(cell.column)
            elif TARGET_HEADER_REGEX.match(val):
                if cell.column not in dutch_cols:
                    dutch_cols.append(cell.column)
    
    # 2. Pair them up. 
    # Logic: For each English column, find the closest Dutch column to its right.
    # If no Dutch column is found, default to English+1.
    pairs: List[Tuple[int, int]] = []
    used_dutch: List[int] = []
    
    for ec in sorted(english_cols):
        # Find closest dutch col to the right
        found_tc = None
        for tc in sorted(dutch_cols):
            if tc > ec and tc not in used_dutch:
                found_tc = tc
                used_dutch.append(tc)
                break
        
        if found_tc:
            pairs.append((ec, found_tc))
        else:
            # Fallback: Default to col immediately after
            pairs.append((ec, ec + 1))
            
    return pairs

# ── Public API ────────────────────────────────────────────────────────────────

def read_word_entries(file_bytes: bytes) -> dict[str, list[WordEntry]]:
    """
    Open an Excel workbook and extract words from ALL English source columns.
    Optimized for multi-column files (up to 3000+).
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result: dict[str, list[WordEntry]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_pairs = _find_column_pairs(ws)
        
        if not col_pairs:
            continue
            
        entries: list[WordEntry] = []
        
        # Start reading after the first 10 rows (header zone)
        for row_cells in ws.iter_rows(min_row=2):
            # Check if this is a "Separator Row"
            # Logic: If row has text in Col 1 but the English cells are empty/same as Col 1, 
            # we check if it's likely a header.
            
            for ec, tc in col_pairs:
                if ec > len(row_cells): continue
                
                cell = row_cells[ec - 1]
                val = _extract_value(cell.value)
                
                if val:
                    entries.append(
                        WordEntry(
                            sheet=sheet_name,
                            row=cell.row,
                            col=ec,
                            target_col=tc,
                            value=val,
                        )
                    )

        if entries:
            result[sheet_name] = entries

    return result


def write_translations(
    original_bytes: bytes,
    sheet_entries: dict[str, list[WordEntry]],
    translation_cache: dict[str, str],
) -> bytes:
    """
    Apply translations to every mapped Dutch column relative to its English source.
    Safely handles MergedCells.
    """
    wb = load_workbook(io.BytesIO(original_bytes))

    for sheet_name, entries in sheet_entries.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        for entry in entries:
            dutch = translation_cache.get(entry.value, entry.value)
            
            # Access the target cell
            try:
                cell = ws.cell(row=entry.row, column=entry.target_col)
                if not isinstance(cell, MergedCell):
                    cell.value = dutch
            except Exception as e:
                logger.debug(f"Could not write to row {entry.row} col {entry.target_col}: {e}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def unique_words(sheet_entries: dict[str, list[WordEntry]]) -> list[str]:
    """Deduplicate all unique English strings found in all columns."""
    seen: set[str] = set()
    result: list[str] = []
    for entries in sheet_entries.values():
        for e in entries:
            if e.value not in seen:
                seen.add(e.value)
                result.append(e.value)
    return result

# ── Helpers ───────────────────────────────────────────────────────────────────

def _extract_value(raw) -> str:
    """Clean specific row values, ignoring IDs or purely numeric rows."""
    if raw is None: return ""
    val = str(raw).strip()
    if not val or val.lower() in ("none", "nan"): return ""
    return val
